"""Generate Excel output with invoice extraction results."""

from __future__ import annotations

from decimal import Decimal
from pathlib import Path
import re

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from app.models.invoice import InvoiceFields

AMOUNT_FIELD_SET = {
    "remuneration",
    "consumption_tax",
    "total_amount",
    "withholding_tax",
}
NUMERIC_FIELD_SET = AMOUNT_FIELD_SET | {"tax_verification"}
TAX_RATE_FACTORS = {
    "0": Decimal("0"),
    "8": Decimal("0.08"),
    "10": Decimal("0.10"),
}


def _excel_value(value):
    if isinstance(value, Decimal):
        if value == value.to_integral_value():
            return int(value)
        return float(value)
    return value


def _format_decimal(value: Decimal) -> str:
    if value == value.to_integral_value():
        return str(int(value))
    return format(value.normalize(), "f")


def _parse_tax_verification_detail(detail: str | None) -> dict[str, Decimal] | None:
    if not detail:
        return None

    matches = re.findall(r"(0|8|10)\s*%?\s*[:：]\s*([+-]?\d[\d,，、]*)\s*円?", detail)
    if not matches:
        return None

    values = {rate: Decimal("0") for rate in TAX_RATE_FACTORS}
    for rate, amount in matches:
        normalized = re.sub(r"[^\d+-]", "", amount)
        if normalized in {"", "+", "-"}:
            continue
        values[rate] = Decimal(normalized)
    return values


def _resolve_tax_verification_value(
    invoice: InvoiceFields,
) -> tuple[int | float | None, str | None]:
    breakdown = _parse_tax_verification_detail(invoice.tax_verification)
    if not breakdown:
        return None, None

    sumproduct_total = sum(
        amount * TAX_RATE_FACTORS[rate]
        for rate, amount in breakdown.items()
    )
    direct_sum_total = sum(breakdown.values(), Decimal("0"))
    expected_tax = invoice.consumption_tax

    mode = "sumproduct"
    resolved = sumproduct_total
    if expected_tax is not None:
        if direct_sum_total == expected_tax and sumproduct_total != expected_tax:
            mode = "sum"
            resolved = direct_sum_total
        elif sumproduct_total == expected_tax:
            mode = "sumproduct"
            resolved = sumproduct_total
        elif abs(direct_sum_total - expected_tax) < abs(sumproduct_total - expected_tax):
            mode = "sum"
            resolved = direct_sum_total

    rate_values = ", ".join(_format_decimal(breakdown[rate]) for rate in ("0", "8", "10"))
    if mode == "sum":
        formula_note = (
            "税率明細は票面税額として判定。"
            f"\n合計 = {_format_decimal(breakdown['0'])} + {_format_decimal(breakdown['8'])} + {_format_decimal(breakdown['10'])}"
            f"\n= {_format_decimal(resolved)}"
        )
    else:
        formula_note = (
            "税率明細は税基として判定。"
            f"\nSUMPRODUCT({{0%, 8%, 10%}}, {{{rate_values}}})"
            f"\n= {_format_decimal(resolved)}"
        )

    return _excel_value(resolved), formula_note

HEADERS = [
    "文件名称",
    "相手先",
    "登録番号",
    "発行日",
    "業務内容",
    "币种",
    "報酬額",
    "消費税額",
    "総額",
    "發票號",
    "消费税核定",
    "源泉税金額",
    "識別モデル",
    "備注",
    "Gemini错误",
    "OpenAI错误",
]

FIELD_MAP = [
    "file_name",
    "issuer",
    "registration_number",
    "issue_date",
    "business_content",
    "currency",
    "remuneration",
    "consumption_tax",
    "total_amount",
    "invoice_number",
    "tax_verification",
    "withholding_tax",
    "source_model",
    "error",
    "gemini_error",
    "openai_error",
]


def write_excel(results: list[InvoiceFields], output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invoice Data"

    # Header styling
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Write headers
    for col_idx, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Write data rows
    for row_idx, invoice in enumerate(results, 2):
        data = invoice.model_dump()
        for col_idx, field in enumerate(FIELD_MAP, 1):
            value = _excel_value(data.get(field))
            if field == "tax_verification":
                value, _ = _resolve_tax_verification_value(invoice)
            cell = ws.cell(
                row=row_idx,
                column=col_idx,
                value="" if value is None else value,
            )
            cell.border = thin_border
            cell.alignment = Alignment(
                vertical="center",
                wrap_text=True,
                horizontal="right" if field in NUMERIC_FIELD_SET else None,
            )
            if field in NUMERIC_FIELD_SET and value is not None:
                cell.number_format = "#,##0.########"

        # Add tax verification detail as comment on 消费税核定 cell
        tax_cell = ws.cell(row=row_idx, column=FIELD_MAP.index("tax_verification") + 1)
        if invoice.tax_verification:
            _, formula_note = _resolve_tax_verification_value(invoice)
            comment_text = f"税率明細:\n{invoice.tax_verification}"
            if formula_note:
                comment_text += f"\n\n計算指引:\n{formula_note}"
            tax_cell.comment = Comment(
                comment_text,
                "Invoice Scanner",
            )

        # Highlight error rows
        if invoice.error:
            error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            for col_idx in range(1, len(HEADERS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = error_fill

    # Auto-adjust column widths
    for col_idx in range(1, len(HEADERS) + 1):
        max_len = len(str(HEADERS[col_idx - 1]))
        for row_idx in range(2, len(results) + 2):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None and val != "":
                max_len = max(max_len, min(len(str(val)), 50))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 4

    # Freeze header row
    ws.freeze_panes = "A2"

    wb.save(str(output_path))
