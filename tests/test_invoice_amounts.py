from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook

from app.models.invoice import InvoiceFields
from app.services.ai_clients.base import AIClient
from app.services.comparator import (
    check_amount_consistency,
    is_amount_consistent,
    maybe_fill_consumption_tax,
    normalize,
)
from app.services.excel_writer import write_excel


def test_invoice_fields_coerce_amounts_to_decimal():
    invoice = InvoiceFields(
        remuneration="¥1,234",
        consumption_tax="Y80",
        total_amount="\\3,860.50",
        withholding_tax="(JPY2960)",
    )

    assert invoice.remuneration == Decimal("1234")
    assert invoice.consumption_tax == Decimal("80")
    assert invoice.total_amount == Decimal("3860.50")
    assert invoice.withholding_tax == Decimal("-2960")


def test_dict_to_invoice_accepts_numeric_json_amounts():
    invoice = AIClient.dict_to_invoice(
        {
            "currency": "JPY",
            "remuneration": 1000,
            "consumption_tax": 100,
            "total_amount": 1100,
            "withholding_tax": 0,
        },
        "sample.jpg",
        "gemini:test",
    )

    assert invoice.remuneration == Decimal("1000")
    assert invoice.consumption_tax == Decimal("100")
    assert invoice.total_amount == Decimal("1100")
    assert invoice.withholding_tax == Decimal("0")


def test_excel_writer_outputs_numeric_cells(tmp_path: Path):
    output_path = tmp_path / "invoice.xlsx"
    write_excel(
        [
            InvoiceFields(
                file_name="sample.jpg",
                remuneration="1000",
                consumption_tax="0",
                total_amount="1100.5",
                withholding_tax="-2960",
            )
        ],
        output_path,
    )

    wb = load_workbook(output_path)
    ws = wb.active
    assert ws["G2"].value == 1000
    assert ws["H2"].value == 0
    assert ws["I2"].value == 1100.5
    assert ws["L2"].value == -2960
    assert ws["G2"].number_format == "#,##0.########"


def test_excel_writer_tax_verification_uses_sumproduct_for_tax_base_detail(tmp_path: Path):
    output_path = tmp_path / "invoice.xlsx"
    write_excel(
        [
            InvoiceFields(
                file_name="sample.pdf",
                consumption_tax="9630",
                tax_verification="0%: 1710円; 10%: 96300円",
            )
        ],
        output_path,
    )

    wb = load_workbook(output_path)
    ws = wb.active
    assert ws["K2"].value == 9630
    assert ws["K2"].number_format == "#,##0.########"
    assert ws["K2"].comment is not None
    assert "SUMPRODUCT" in ws["K2"].comment.text


def test_excel_writer_tax_verification_stays_numeric_for_receipt_tax_amount_detail(tmp_path: Path):
    output_path = tmp_path / "invoice.xlsx"
    write_excel(
        [
            InvoiceFields(
                file_name="receipt.jpg",
                consumption_tax="92",
                tax_verification="8%: 92円",
            )
        ],
        output_path,
    )

    wb = load_workbook(output_path)
    ws = wb.active
    assert ws["K2"].value == 92
    assert ws["K2"].comment is not None
    assert "票面税額" in ws["K2"].comment.text


def test_normalize_supports_numeric_values():
    assert normalize(Decimal("3860")) == "3860"
    assert normalize(3860) == "3860"
    assert normalize("3,860円") == "3860"


def test_tax_verification_is_normalized_to_all_rates():
    invoice = InvoiceFields(tax_verification="10%: 96300円; 0%: 1710円")

    assert invoice.tax_verification == "0%: 1710円; 8%: 0円; 10%: 96300円"


# --- Amount consistency tests ---


def test_is_amount_consistent_with_matching_amounts():
    invoice = InvoiceFields(
        remuneration="100000", consumption_tax="10000", total_amount="110000"
    )
    assert is_amount_consistent(invoice) is True


def test_is_amount_consistent_with_mismatched_amounts():
    invoice = InvoiceFields(
        remuneration="100000", consumption_tax="10000", total_amount="120000"
    )
    assert is_amount_consistent(invoice) is False


def test_is_amount_consistent_returns_false_when_fields_missing():
    invoice = InvoiceFields(remuneration="100000", total_amount="110000")
    assert is_amount_consistent(invoice) is False


def test_check_amount_consistency_no_warnings_when_consistent():
    invoice = InvoiceFields(
        remuneration="100000", consumption_tax="10000", total_amount="110000"
    )
    assert check_amount_consistency(invoice) == []


def test_check_amount_consistency_warns_on_mismatch():
    invoice = InvoiceFields(
        remuneration="100000", consumption_tax="10000", total_amount="99790"
    )
    warnings = check_amount_consistency(invoice)
    assert len(warnings) == 1
    assert "金額不整合" in warnings[0]


def test_check_amount_consistency_flags_withholding_deducted_total():
    # 総額 should be pre-withholding (報酬+税=110000), not 差引支払額 (99790)
    # If 総額 = 報酬 + 税 - 源泉, it means 差引支払額 was mistaken for 総額
    invoice = InvoiceFields(
        remuneration="100000",
        consumption_tax="10000",
        total_amount="99790",
        withholding_tax="-10210",
    )
    warnings = check_amount_consistency(invoice)
    assert len(warnings) == 1
    assert "差引支払額" in warnings[0]


def test_check_amount_consistency_warns_on_abnormal_tax_rate():
    # 消費税 / 報酬額 = 20% which is abnormal
    invoice = InvoiceFields(
        remuneration="100000", consumption_tax="20000", total_amount="120000"
    )
    warnings = check_amount_consistency(invoice)
    assert any("消費税率が異常" in w for w in warnings)


def test_check_amount_consistency_no_tax_rate_warning_for_valid_rates():
    # 10% tax rate is valid
    invoice = InvoiceFields(
        remuneration="100000", consumption_tax="10000", total_amount="110000"
    )
    assert not any("消費税率" in w for w in check_amount_consistency(invoice))

    # 8% tax rate is valid
    invoice_8 = InvoiceFields(
        remuneration="100000", consumption_tax="8000", total_amount="108000"
    )
    assert not any("消費税率" in w for w in check_amount_consistency(invoice_8))


def test_excel_writer_tax_verification_mismatch_warning(tmp_path: Path):
    output_path = tmp_path / "invoice.xlsx"
    # Tax verification breakdown doesn't match consumption_tax at all
    write_excel(
        [
            InvoiceFields(
                file_name="mismatch.pdf",
                consumption_tax="5000",
                tax_verification="0%: 0円; 8%: 1000円; 10%: 2000円",
            )
        ],
        output_path,
    )

    wb = load_workbook(output_path)
    ws = wb.active
    assert ws["K2"].comment is not None
    assert "一致しません" in ws["K2"].comment.text


# --- consumption_tax backfill from tax_verification ---


def test_fill_consumption_tax_from_tax_base_verification():
    # tax_verification has tax bases → SUMPRODUCT gives consumption_tax
    invoice = InvoiceFields(
        file_name="test.pdf",
        remuneration="100000",
        total_amount="110000",
        tax_verification="0%: 0円; 8%: 0円; 10%: 100000円",
    )
    assert invoice.consumption_tax is None
    filled = maybe_fill_consumption_tax(invoice)
    assert filled is True
    assert invoice.consumption_tax == Decimal("10000")


def test_fill_consumption_tax_from_receipt_tax_amount_verification():
    # tax_verification has direct tax amounts → direct sum gives consumption_tax
    invoice = InvoiceFields(
        file_name="receipt.jpg",
        remuneration="1150",
        total_amount="1242",
        tax_verification="0%: 0円; 8%: 92円; 10%: 0円",
    )
    assert invoice.consumption_tax is None
    filled = maybe_fill_consumption_tax(invoice)
    assert filled is True
    assert invoice.consumption_tax == Decimal("92")


def test_fill_consumption_tax_skips_when_already_present():
    invoice = InvoiceFields(
        file_name="test.pdf",
        consumption_tax="10000",
        tax_verification="0%: 0円; 8%: 0円; 10%: 100000円",
    )
    filled = maybe_fill_consumption_tax(invoice)
    assert filled is False
    assert invoice.consumption_tax == Decimal("10000")


def test_fill_consumption_tax_skips_without_verification():
    invoice = InvoiceFields(file_name="test.pdf")
    filled = maybe_fill_consumption_tax(invoice)
    assert filled is False
    assert invoice.consumption_tax is None
